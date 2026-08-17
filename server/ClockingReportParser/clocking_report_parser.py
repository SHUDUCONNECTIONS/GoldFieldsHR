"""
Clocking Report Parser
=======================
Parses a "South Deep / Gold Fields - Individual Clocking History" PDF report
(as produced by the web_clock_report_individual program) and exports it to a
formatted Excel workbook containing a single "Timesheet" sheet: Name/Employee
Number/Occupation, then one row per calendar day (Day/Date/Shift/1st/2nd/
Hrs of work/Planned/O-T Minutes/S-T Minutes/Comments), a totals row, and a
Planned/Actual/Overtime summary box - see build_timesheet() below for exactly
how each column is derived.

USAGE
-----
    pip install pdfplumber pandas openpyxl

    Single file:
        python clocking_report_parser.py path/to/report.pdf [output.xlsx]

    Batch (every *.pdf in a folder):
        python clocking_report_parser.py path/to/folder [output_folder]

If the .xlsx / output folder is omitted, each workbook is written as
"<pdf name>_parsed.xlsx" next to its input PDF. In batch mode, one failing
PDF is logged and skipped rather than stopping the whole run.

Add --work-days mon,tue,wed,thu,fri and/or --hours-per-day 8 to match a
company's actual schedule instead of the Mon-Fri/8h default - see the
"PLANNED HOURS" note below for why this is a guess either way.

PLANNED HOURS
-------------
There is no roster in the source PDF, so "Shift" and "Planned" are guessed
purely from which days of the week --work-days lists as scheduled working
days - not from what any individual employee actually works. This is
configurable (per run, not per employee), but it's still a guess: someone
on a non-standard week (night shift, rotating roster, part-time) will get
a wrong Planned figure regardless of how --work-days is set.

HOW THE PARSING WORKS
----------------------
This is a *layout parser*, not an NLP/ML system - the report is a fixed-column
table, so the most reliable approach is to read each PDF word's x/y position
(via pdfplumber) and bucket words into columns by their x-coordinate, rather
than trying to regex the flattened text (which loses the IN vs OUT column
distinction whenever both are printed on the same line).

Column x-position anchors (points from left edge of page), taken from the
report's own header row:
    IN  columns : Date=25  Time=125  From=175  To=225  Point=275  Type=325
    OUT columns : Time=525 From=575  To=625    Point=675 Type=725

HOURS WORKED METHODOLOGY
--------------------------
Rather than using each calendar day's first-to-last clocking (which breaks
down for overnight shifts and for people working several nights in a row
with short rest gaps), this script pairs each "Work" type clocking through
Point "U01" (the underground workplace access point): the IN event marks
entry to the workplace and the following OUT event marks exit. The
difference between them is that shift's "hours worked". This deliberately
excludes surface travel time through access gates (FCP/SG1/SG2/A01/A11/etc.)
before and after the shift.

If your report uses a different "work area" point code, change WORK_POINT
below.
"""

import argparse
import json
import os
import re
import sys
import tempfile
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
import pdfplumber
import pikepdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------

# The point code that represents the actual underground workplace / work area.
# Used to compute "Hours Worked" from IN/OUT pairs of Type == "Work".
WORK_POINT = "U01"

# Gap threshold (hours) used only as a sanity fallback if no Work/U01 pairs
# are found in the report (see build_hours_worked_fallback()).
FALLBACK_SESSION_GAP_HOURS = 14

# x-coordinate boundaries used to classify each word into IN vs OUT columns,
# and then into a specific field within that side. Adjust these if your PDF's
# layout differs (run inspect_columns() below to check).
IN_OUT_SPLIT_X = 450

IN_FIELD_BOUNDS = [
    (75, "Date"),
    (150, "Time"),
    (200, "From"),
    (250, "To"),
    (300, "Point"),
    (float("inf"), "Type"),
]

OUT_FIELD_BOUNDS = [
    (550, "Time"),
    (600, "From"),
    (650, "To"),
    (700, "Point"),
    (float("inf"), "Type"),
]

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
TIME_RE = re.compile(r"^\d{2}:\d{2}(:\d{2})?$")

SKIP_LINE_SUBSTRINGS = [
    "Program - web_clock_report_individual",
    "South Deep",
    "Date From",
    "Date To",
    "Individual Clocking History",
    "Ind Nbr",
    "SurName",
    "First Names",
    "IN CLOCKINGS",
    "OUT CLOCKINGS",
    "End of Clocking Report",
]


# ----------------------------------------------------------------------------
# Step 1: Parse the PDF into structured records
# ----------------------------------------------------------------------------

def classify_field(x0, bounds):
    for boundary, name in bounds:
        if x0 < boundary:
            return name
    return bounds[-1][1]


def chunk_into_fields(word_texts):
    """Given a flat list of word strings for one side (IN or OUT) of a line,
    split them into groups of 5: Time, From, To, Point, Type."""
    records = []
    i = 0
    while i < len(word_texts):
        if i + 5 <= len(word_texts) and TIME_RE.match(word_texts[i]):
            records.append(
                {
                    "Time": word_texts[i],
                    "From": word_texts[i + 1],
                    "To": word_texts[i + 2],
                    "Point": word_texts[i + 3],
                    "Type": word_texts[i + 4],
                }
            )
            i += 5
        else:
            i += 1
    return records


def parse_metadata(first_page_text):
    meta = {}
    for line in first_page_text.split("\n"):
        if line.startswith("Ind Nbr"):
            m = re.search(r"Ind Nbr:\s*(\S+)\s*Org Unit:\s*(.+)", line)
            if m:
                meta["Ind Nbr"] = m.group(1)
                meta["Org Unit"] = m.group(2).strip()
        if line.startswith("SurName"):
            m = re.search(r"SurName:\s*(\S+)\s*Designation:\s*(.+)", line)
            if m:
                meta["SurName"] = m.group(1)
                meta["Designation"] = m.group(2).strip()
        if line.startswith("First Names"):
            m = re.search(
                r"First Names:\s*(.+?)\s*Type:\s*(\S+)\s*Environment:\s*(\S+)", line
            )
            if m:
                meta["First Names"] = m.group(1).strip()
                meta["Type"] = m.group(2)
                meta["Environment"] = m.group(3)
        if "Date From" in line:
            m = re.search(r"Date From\s*:\s*(\S+)", line)
            if m:
                meta["Date From"] = m.group(1)
        if "Date To" in line:
            m = re.search(r"Date To\s*:\s*(\S+)", line)
            if m:
                meta["Date To"] = m.group(1)
    return meta


def _open_pdf(pdf_path):
    """Open pdf_path with pdfplumber, transparently repairing malformed PDFs
    via pikepdf/qpdf if the first attempt fails.

    Some scanners/exporters write PDFs with a broken or missing xref/trailer
    (pdfminer then raises e.g. "No /Root object! - Is this really a PDF?")
    that Adobe/Chrome silently repair on open. pikepdf wraps qpdf, which can
    rebuild the xref table and recover the same file those viewers show.
    """
    try:
        return pdfplumber.open(pdf_path)
    except Exception as original_error:
        try:
            fd, repaired_path = tempfile.mkstemp(suffix=".pdf")
            os.close(fd)
            with pikepdf.open(pdf_path) as pdf:
                pdf.save(repaired_path)
            return pdfplumber.open(repaired_path)
        except Exception:
            raise original_error


def parse_pdf(pdf_path):
    """Returns (meta_dict, list_of_clocking_record_dicts).

    Raises ValueError if no clocking records are found. The column x-positions
    and HH:MM[:SS] row format this relies on (see module docstring) were
    tuned against one report template - other exports may lay out columns or
    format times differently. Rather than guess, the error includes a preview
    of what was actually extracted, so a real mismatch can be diagnosed from
    the error message alone instead of needing the source PDF.
    """
    records = []
    current_date = None
    any_words = False
    sample_lines = []

    with _open_pdf(pdf_path) as pdf:
        meta = parse_metadata(pdf.pages[0].extract_text())

        for page in pdf.pages:
            words = page.extract_words()
            if words:
                any_words = True

            # Group words into visual lines by their y-position ("top").
            lines = {}
            for w in words:
                top = round(w["top"])
                lines.setdefault(top, []).append(w)

            for top in sorted(lines.keys()):
                ws = sorted(lines[top], key=lambda w: w["x0"])
                line_str = " ".join(w["text"] for w in ws)

                if any(s in line_str for s in SKIP_LINE_SUBSTRINGS):
                    continue
                if line_str.strip() == "Date Time From To Point Type Time From To Point Type":
                    continue
                if not line_str.strip():
                    continue

                if len(sample_lines) < 4:
                    sample_lines.append(line_str[:120])

                in_words = [w for w in ws if w["x0"] < IN_OUT_SPLIT_X]
                out_words = [w for w in ws if w["x0"] >= IN_OUT_SPLIT_X]

                # A date at the start of the IN side updates current_date.
                line_date = current_date
                if in_words and DATE_RE.match(in_words[0]["text"]):
                    line_date = in_words[0]["text"]
                    current_date = line_date
                    in_words = in_words[1:]

                in_texts = [w["text"] for w in in_words]
                out_texts = [w["text"] for w in out_words]

                for rec in chunk_into_fields(in_texts):
                    records.append({"Date": line_date, "Direction": "IN", **rec})
                for rec in chunk_into_fields(out_texts):
                    records.append({"Date": line_date, "Direction": "OUT", **rec})

    if not records:
        if not any_words:
            raise ValueError(
                "No clocking records found - the PDF has no extractable text "
                "(likely a scanned image rather than a generated report)."
            )
        if not sample_lines:
            raise ValueError(
                "No clocking records found - every line matched a known header/"
                "boilerplate pattern, so the report may be empty for this date range."
            )
        preview = "  |  ".join(sample_lines)
        raise ValueError(
            "No clocking records found - none of the report's lines matched the "
            "expected 'HH:MM[:SS] From To Point Type' row format (see module "
            f"docstring for the column layout this parser expects). Lines seen: {preview}"
        )

    return meta, records


# ----------------------------------------------------------------------------
# Step 2: Build derived tables (daily summary, hours worked, weekly hours)
# ----------------------------------------------------------------------------

def build_dataframe(records):
    df = pd.DataFrame(records)
    df["Datetime"] = pd.to_datetime(df["Date"] + " " + df["Time"])
    df = df.sort_values("Datetime").reset_index(drop=True)
    return df


def build_daily_summary(df, date_from, date_to):
    grp = df.groupby("Date")
    daily = grp.agg(
        First_Clocking=("Time", "min"),
        Last_Clocking=("Time", "max"),
        Num_Events=("Time", "count"),
    ).reset_index()

    start = datetime.strptime(date_from, "%Y-%m-%d")
    end = datetime.strptime(date_to, "%Y-%m-%d")
    all_dates = []
    cur = start
    while cur <= end:
        all_dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)

    full_daily = pd.DataFrame({"Date": all_dates}).merge(daily, on="Date", how="left")
    full_daily["Num_Events"] = full_daily["Num_Events"].fillna(0).astype(int)
    full_daily["First_Clocking"] = full_daily["First_Clocking"].fillna("")
    full_daily["Last_Clocking"] = full_daily["Last_Clocking"].fillna("")
    full_daily["Status"] = full_daily["Num_Events"].apply(
        lambda n: "No Clocking" if n == 0 else "Clocked"
    )
    return full_daily


def build_hours_worked(df, work_point=WORK_POINT):
    """Pair IN/OUT clockings of Type == 'Work' at `work_point` into shifts."""
    work = df[(df["Type"] == "Work") & (df["Point"] == work_point)].reset_index(drop=True)

    if work.empty:
        return build_hours_worked_fallback(df)

    shifts = []
    i = 0
    while i < len(work):
        a = work.iloc[i]
        if a["Direction"] == "IN" and i + 1 < len(work) and work.iloc[i + 1]["Direction"] == "OUT":
            b = work.iloc[i + 1]
            dur_h = (b["Datetime"] - a["Datetime"]).total_seconds() / 3600
            shifts.append(
                {
                    "Shift Date": a["Date"],
                    "Clock In": a["Time"],
                    "Clock Out Date": b["Date"],
                    "Clock Out": b["Time"],
                    "Hours Worked": round(dur_h, 2),
                    "Note": "Short / possible anomaly" if dur_h < 1 else "",
                }
            )
            i += 2
        else:
            shifts.append(
                {
                    "Shift Date": a["Date"],
                    "Clock In": a["Time"] if a["Direction"] == "IN" else "",
                    "Clock Out Date": a["Date"] if a["Direction"] == "OUT" else "",
                    "Clock Out": a["Time"] if a["Direction"] == "OUT" else "",
                    "Hours Worked": 0,
                    "Note": "Unmatched clocking - could not pair",
                }
            )
            i += 1

    return pd.DataFrame(shifts)


def build_hours_worked_fallback(df, gap_hours=FALLBACK_SESSION_GAP_HOURS):
    """Used only if no Work/`WORK_POINT` events exist in the report: clusters
    ALL clockings into sessions separated by gaps > gap_hours. Less accurate
    than the Work-point pairing (may merge consecutive shifts with short
    rest gaps), so only used as a last resort."""
    gap = df["Datetime"].diff().dt.total_seconds() / 3600
    new_session = (gap > gap_hours) | gap.isna()
    df = df.copy()
    df["Session"] = new_session.cumsum()

    sessions = df.groupby("Session").agg(Start=("Datetime", "min"), End=("Datetime", "max"))
    sessions["Hours Worked"] = (
        (sessions["End"] - sessions["Start"]).dt.total_seconds() / 3600
    ).round(2)
    sessions = sessions.reset_index()

    return pd.DataFrame(
        {
            "Shift Date": sessions["Start"].dt.strftime("%Y-%m-%d"),
            "Clock In": sessions["Start"].dt.strftime("%H:%M:%S"),
            "Clock Out Date": sessions["End"].dt.strftime("%Y-%m-%d"),
            "Clock Out": sessions["End"].dt.strftime("%H:%M:%S"),
            "Hours Worked": sessions["Hours Worked"],
            "Note": "Fallback: clustered by gap > "
            f"{gap_hours}h (no Work/{WORK_POINT} events found)",
        }
    )


WEEKDAY_ABBR = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# Defaults used when the caller doesn't configure a schedule: Mon-Fri, 8h/day.
DEFAULT_WORK_DAYS = frozenset({0, 1, 2, 3, 4})
DEFAULT_HOURS_PER_DAY = 8.0

# Payroll multipliers applied to O/T Minutes and S/T Minutes to get paid
# overtime hours (see write_timesheet_sheet's "Overtime Pay" summary line).
OT_RATE = 1.5
ST_RATE = 2.0


def parse_work_days(spec):
    """Parses a comma-separated weekday spec like "mon,tue,wed,thu,fri" into
    a set of Python weekday() indices (Monday=0 ... Sunday=6)."""
    days = set()
    for part in spec.split(","):
        part = part.strip()[:3].lower()
        if part not in WEEKDAY_ABBR:
            raise ValueError(
                f"Unknown day '{part}' - use a comma-separated list of "
                "mon/tue/wed/thu/fri/sat/sun"
            )
        days.add(WEEKDAY_ABBR[part])
    return days


def describe_schedule(work_days, hours_per_day):
    """Human-readable summary of the configured schedule, e.g.
    "Mon-Fri, 8.0h/day" or "Mon, Wed, Fri, 10.0h/day" for non-contiguous sets."""
    ordered = sorted(work_days)
    names = [WEEKDAY_NAMES[d][:3] for d in ordered]
    is_contiguous = ordered == list(range(ordered[0], ordered[0] + len(ordered))) if ordered else False
    span = f"{names[0]}-{names[-1]}" if is_contiguous and len(names) > 1 else ", ".join(names)
    return f"{span or 'no scheduled days'}, {hours_per_day:g}h/day"


def build_timesheet(df, meta, work_days=None, hours_per_day=DEFAULT_HOURS_PER_DAY, rotating=False):
    """Builds the day-by-day timesheet table (Day/Date/Shift/1st/2nd/Hrs of
    work/Planned/O-T Minutes/S-T Minutes/Comments) shown in the target report
    layout. One row per calendar day in the report's date range.

    Shift/Planned are assigned one of two ways, since there's no roster in
    the source PDF either way:

    - `rotating=False` (default): by day-of-week. `work_days` (a set of
      weekday() indices, Monday=0..Sunday=6; default Mon-Fri) get "D/S" with
      `hours_per_day` planned hours, every other day gets "OFF" with no
      planned hours. Configure `work_days`/`hours_per_day` to match the
      company's normal schedule if it isn't a plain Mon-Fri week.
    - `rotating=True`: for shifts that don't follow a fixed weekly pattern
      (e.g. 4-on/4-off) - `work_days` is ignored, and instead every day with
      at least one clocking is treated as a scheduled `hours_per_day` day
      (every day with no clocking is OFF). This can't tell a genuine rest
      day from a missed clocking, but a day-of-week guess would be no more
      accurate for a rotation that isn't tied to the calendar week.

    1st/2nd are the first and last clocking of the day (any type), and
    Hrs of work is simply their difference, matching the reference layout.

    O/T Minutes (1.5x) is any time worked beyond `hours_per_day` on a
    non-Sunday - e.g. staying past a planned 8h shift. S/T Minutes (2.0x,
    "Sunday Time") is ALL time worked on a Sunday, regardless of whether
    Sunday is a scheduled day or how many hours were planned - Sunday work
    never counts toward O/T Minutes.
    """
    if not rotating and work_days is None:
        work_days = DEFAULT_WORK_DAYS

    start = datetime.strptime(meta["Date From"], "%Y-%m-%d")
    end = datetime.strptime(meta["Date To"], "%Y-%m-%d")

    span = df.groupby("Date")["Datetime"].agg(First="min", Last="max")

    rows = []
    cur = start
    while cur <= end:
        date_str = cur.strftime("%Y-%m-%d")
        is_scheduled = date_str in span.index if rotating else cur.weekday() in work_days

        if date_str in span.index:
            first_dt = span.loc[date_str, "First"]
            last_dt = span.loc[date_str, "Last"]
            hrs_of_work = last_dt - first_dt
        else:
            first_dt = last_dt = hrs_of_work = None

        comment = ""
        if is_scheduled and first_dt is None:
            comment = "No clocking recorded"
        elif not is_scheduled and first_dt is not None:
            comment = "Worked on OFF day"

        is_sunday = cur.weekday() == WEEKDAY_ABBR["sun"]
        planned = timedelta(hours=hours_per_day) if is_scheduled else None

        if hrs_of_work is None:
            ot_minutes = st_minutes = None
        elif is_sunday:
            ot_minutes, st_minutes = None, hrs_of_work
        else:
            excess = hrs_of_work - (planned or timedelta())
            ot_minutes = excess if excess > timedelta() else None
            st_minutes = None

        rows.append(
            {
                "Day": cur.strftime("%A"),
                "Date": cur,
                "Shift": "D/S" if is_scheduled else "OFF",
                "1st": first_dt.time() if first_dt is not None else None,
                "2nd": last_dt.time() if last_dt is not None else None,
                "Hrs of work": hrs_of_work,
                "Planned": planned,
                "O/T Minutes": ot_minutes,
                "S/T Minutes": st_minutes,
                "Comments": comment,
                "Is Off": not is_scheduled,
            }
        )
        cur += timedelta(days=1)

    return pd.DataFrame(rows)


# ----------------------------------------------------------------------------
# Step 3: Write the Excel workbook
# ----------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=14, bold=True)
LABEL_FONT = Font(name="Arial", size=11, bold=True)
INFO_FONT = Font(name="Arial", size=11)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
VALUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def style_header_row(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def write_timesheet_sheet(
    wb, meta, timesheet_df, work_days=None, hours_per_day=DEFAULT_HOURS_PER_DAY, rotating=False
):
    ws = wb.active
    ws.title = "Timesheet"

    name = f"{meta.get('First Names', '')} {meta.get('SurName', '')}".strip()

    ws["A1"] = "Individual Clocking History - Timesheet"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A3"] = "Name"
    ws["A3"].font = LABEL_FONT
    ws["B3"] = name
    ws["B3"].font = INFO_FONT
    ws.merge_cells("B3:D3")

    ws["A4"] = "Employee Number"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = meta.get("Ind Nbr", "")
    ws["B4"].font = INFO_FONT
    ws["B4"].fill = VALUE_FILL
    ws.merge_cells("B4:D4")

    ws["F3"] = "Occupation"
    ws["F3"].font = LABEL_FONT
    ws["G3"] = meta.get("Designation", "")
    ws["G3"].font = INFO_FONT
    ws["G3"].fill = VALUE_FILL
    ws.merge_cells("G3:J3")

    headers = [
        "Day", "Date", "Shift", "1st", "2nd", "Hrs of work", "Planned",
        "O/T Minutes", "S/T Minutes", "Comments",
    ]
    header_row = 6
    style_header_row(ws, headers, row=header_row)

    first_data_row = header_row + 1
    for i, row in timesheet_df.iterrows():
        r = first_data_row + i
        vals = [
            row["Day"], row["Date"], row["Shift"], row["1st"], row["2nd"],
            row["Hrs of work"], row["Planned"], row["O/T Minutes"],
            row["S/T Minutes"], row["Comments"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            header_name = headers[c - 1]
            if header_name == "Date":
                cell.number_format = "yyyy/mm/dd"
            elif header_name in ("1st", "2nd"):
                cell.number_format = "h:mm:ss AM/PM"
                cell.alignment = Alignment(horizontal="center")
            elif header_name in ("Hrs of work", "Planned", "O/T Minutes", "S/T Minutes"):
                cell.number_format = "[h]:mm"
                cell.alignment = Alignment(horizontal="center")
            elif header_name == "Shift":
                cell.alignment = Alignment(horizontal="center")
        if row["Is Off"]:
            ws.cell(row=r, column=3).fill = WEEKEND_FILL
        if row["Comments"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = FLAG_FILL

    # Written as live Excel formulas (not Python-computed values) so that
    # editing any daily row - e.g. correcting a mis-parsed "Hrs of work" -
    # recalculates the TOTAL row and the summary box below it automatically.
    # build_workbook() also patches in a cached result for each formula (see
    # _write_cached_formula_values) so they're correct on first open too,
    # not just after an edit - openpyxl has no calc engine, so a formula
    # with no cached value shows blank/0 until Excel recalculates it, which
    # it defers while the file is still in Protected View.
    def sum_timedeltas(values):
        # pandas stores this column as timedelta64[ns], so missing entries
        # are NaT rather than None - and bool(NaT) is True, so a plain
        # `if v` truthiness check lets NaT through and poisons the sum
        # (any arithmetic with NaT produces NaT). pd.notna() is required.
        total = timedelta()
        for v in values:
            if pd.notna(v):
                total += v
        return total

    hrs_total = sum_timedeltas(timesheet_df["Hrs of work"])
    planned_total = sum_timedeltas(timesheet_df["Planned"])
    ot_total = sum_timedeltas(timesheet_df["O/T Minutes"])
    st_total = sum_timedeltas(timesheet_df["S/T Minutes"])

    total_row = first_data_row + len(timesheet_df)
    last_data_row = total_row - 1
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    total_formulas = {
        6: (f"SUM(F{first_data_row}:F{last_data_row})", hrs_total),
        7: (f"SUM(G{first_data_row}:G{last_data_row})", planned_total),
        8: (f"SUM(H{first_data_row}:H{last_data_row})", ot_total),
        9: (f"SUM(I{first_data_row}:I{last_data_row})", st_total),
    }
    cached_values = {}
    for col, (formula, cached_total) in total_formulas.items():
        cell = ws.cell(row=total_row, column=col, value=f"={formula}")
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.number_format = "[h]:mm"
        cell.alignment = Alignment(horizontal="center")
        cached_values[cell.coordinate] = cached_total.total_seconds() / 86400
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).border = BORDER

    planned_cell, actual_cell = f"G{total_row}", f"F{total_row}"
    ot_cell, st_cell = f"H{total_row}", f"I{total_row}"

    box_row = total_row + 3
    ws.cell(row=box_row, column=6, value="Planned Hours").font = LABEL_FONT
    c = ws.cell(row=box_row, column=7, value=f"={planned_cell}")
    c.number_format = "[h]:mm"
    cached_values[c.coordinate] = planned_total.total_seconds() / 86400

    ws.cell(row=box_row + 1, column=6, value="Actual Hours").font = LABEL_FONT
    c = ws.cell(row=box_row + 1, column=7, value=f"={actual_cell}")
    c.number_format = "[h]:mm"
    cached_values[c.coordinate] = hrs_total.total_seconds() / 86400

    ws.cell(row=box_row + 2, column=6, value="Overtime Hours").font = LABEL_FONT
    # O/T Minutes and S/T Minutes cells hold Excel durations (a fraction of a
    # day) - Overtime = sum of the daily O/T + S/T columns, not Actual -
    # Planned (a net figure would let an under-worked day cancel out
    # overtime earned on another day, which isn't how overtime pay works),
    # multiplied by 24 to turn the day-fraction into a decimal hour count.
    c = ws.cell(row=box_row + 2, column=7, value=f"=({ot_cell}+{st_cell})*24")
    c.number_format = '0.00 "h"'
    overtime_hours = (ot_total + st_total).total_seconds() / 3600
    cached_values[c.coordinate] = overtime_hours

    # Paid overtime hours - O/T and S/T are worked at different rates, so
    # they're weighted separately rather than multiplying the flat total.
    ws.cell(row=box_row + 3, column=6, value="Overtime Pay (hours)").font = LABEL_FONT
    c = ws.cell(
        row=box_row + 3, column=7,
        value=f"=({ot_cell}*24*{OT_RATE})+({st_cell}*24*{ST_RATE})",
    )
    c.number_format = '0.00 "h"'
    ot_hours = ot_total.total_seconds() / 3600
    st_hours = st_total.total_seconds() / 3600
    cached_values[c.coordinate] = ot_hours * OT_RATE + st_hours * ST_RATE

    if rotating:
        schedule_note = (
            f"'Shift'/'Planned' follow a rotating {hours_per_day:g}h/day shift (e.g. "
            "4-on/4-off) - any day with a clocking is treated as scheduled, since a "
            "rotation isn't tied to the calendar week and there's no roster to check "
            "against."
        )
    else:
        schedule_desc = describe_schedule(work_days or DEFAULT_WORK_DAYS, hours_per_day)
        schedule_note = (
            f"'Shift'/'Planned' are assigned by day-of-week ({schedule_desc}, all "
            "other days = OFF), not an actual roster."
        )
    note_row = box_row + 5
    note = ws.cell(
        row=note_row, column=1,
        value=(
            "Note: Overtime Hours = O/T Minutes (time worked beyond the planned daily "
            "hours) + S/T Minutes (all time worked on a Sunday). Overtime Pay (hours) = "
            f"O/T Minutes x {OT_RATE:g} + S/T Minutes x {ST_RATE:g}. " + schedule_note
        ),
    )
    note.font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 30

    ws.freeze_panes = f"A{first_data_row}"
    # Widths are sized for the widest of: the header text + its autofilter
    # dropdown arrow, the data it holds, and (for A/F) the row-3/4 and
    # summary-box labels that share those columns ("Employee Number",
    # "Overtime Pay (hours)", etc.) - not just the header text alone.
    for c, w in enumerate([20, 12, 11, 12, 12, 24, 12, 16, 16, 26], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = f"A{header_row}:J{total_row - 1}"

    return cached_values


_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _write_cached_formula_values(xlsx_path, cached_values):
    """Patch a cached <v> result into formula cells after wb.save().

    openpyxl has no calculation engine, so it saves a formula cell as
    <c r="F96"><f>SUM(...)</f><v></v></c> - an empty cached result. Excel
    treats that as needing recalculation before it can display anything,
    which it defers while the file is still in Protected View (the default
    state for anything downloaded via a browser), so the totals/summary box
    show blank until the user clicks "Enable Editing". A genuine
    Excel-authored file always carries a cached result alongside the
    formula; this fills that in with the same number the formula itself
    computes, so it's correct on first open. The formula string is left
    untouched, so it still recalculates normally the moment a dependent
    cell is edited.

    `cached_values` is {cell_ref: number}, e.g. {"F96": 0.40277...}.
    """
    ET.register_namespace("", _XLSX_MAIN_NS)
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        contents = {name: zin.read(name) for name in zin.namelist()}

    sheet_name = next(n for n in contents if n.startswith("xl/worksheets/sheet"))
    root = ET.fromstring(contents[sheet_name])
    for c in root.iter(f"{{{_XLSX_MAIN_NS}}}c"):
        ref = c.get("r")
        if ref not in cached_values:
            continue
        v = c.find(f"{{{_XLSX_MAIN_NS}}}v")
        if v is None:
            v = ET.SubElement(c, f"{{{_XLSX_MAIN_NS}}}v")
        v.text = repr(float(cached_values[ref]))
    contents[sheet_name] = ET.tostring(root, encoding="UTF-8", xml_declaration=True)

    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in contents.items():
            zout.writestr(name, data)


def build_workbook(
    meta, timesheet_df, out_path, work_days=None, hours_per_day=DEFAULT_HOURS_PER_DAY, rotating=False
):
    wb = Workbook()
    cached_values = write_timesheet_sheet(
        wb, meta, timesheet_df, work_days=work_days, hours_per_day=hours_per_day, rotating=rotating
    )
    # Formulas still recalculate normally on edit (see write_timesheet_sheet
    # / _write_cached_formula_values) - fullCalcOnLoad is a best-effort
    # extra nudge for viewers that do run a full recalc on open.
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)
    _write_cached_formula_values(out_path, cached_values)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def process_one(
    pdf_path, out_path, log=print, work_days=None, hours_per_day=DEFAULT_HOURS_PER_DAY, rotating=False
):
    """Parses a single Individual Clocking History PDF and writes its workbook.
    Raises on failure (caller decides whether to keep going in a batch).

    `work_days`/`hours_per_day`/`rotating` configure the guessed Shift/Planned
    schedule (see build_timesheet) - default Mon-Fri, 8h/day, not rotating."""
    log(f"Parsing {pdf_path} ...")
    meta, records = parse_pdf(str(pdf_path))

    df = build_dataframe(records)
    full_daily = build_daily_summary(df, meta["Date From"], meta["Date To"])
    shifts_df = build_hours_worked(df)
    timesheet_df = build_timesheet(
        df, meta, work_days=work_days, hours_per_day=hours_per_day, rotating=rotating
    )

    build_workbook(
        meta, timesheet_df, out_path, work_days=work_days, hours_per_day=hours_per_day, rotating=rotating
    )

    total_hours = shifts_df["Hours Worked"].sum()
    log(f"Parsed {len(records)} clocking events across {full_daily.shape[0]} calendar days.")
    log(f"Identified {len(shifts_df)} shift(s), total hours worked: {total_hours:.2f}")
    log(f"Workbook written to: {out_path}")


def run_batch(
    input_paths,
    output_dir=None,
    log=print,
    work_days=None,
    hours_per_day=DEFAULT_HOURS_PER_DAY,
    rotating=False,
):
    """Shared entry point for both the CLI and the GUI app.

    `input_paths` is a list of files and/or folders (folders are expanded to
    every *.pdf inside them). Each PDF is parsed independently - one failure
    is logged and skipped rather than aborting the rest.

    `work_days`/`hours_per_day`/`rotating` configure the guessed Shift/Planned
    schedule (see build_timesheet) - default Mon-Fri, 8h/day, not rotating.

    Returns (ok, failed) where `ok` is a list of (pdf_path, out_path) and
    `failed` is a list of (pdf_path, error_message).
    """
    pdf_files = []
    for p in input_paths:
        p = Path(p)
        if p.is_dir():
            pdf_files.extend(sorted(p.glob("*.pdf")))
        elif p.suffix.lower() == ".pdf":
            pdf_files.append(p)

    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    ok, failed = [], []
    for f in pdf_files:
        out_path = (output_dir / (f.stem + "_parsed.xlsx")) if output_dir else f.with_name(
            f.stem + "_parsed.xlsx"
        )
        try:
            process_one(
                f, out_path, log=log, work_days=work_days, hours_per_day=hours_per_day, rotating=rotating
            )
            ok.append((f, out_path))
        except Exception as e:
            log(f"  FAILED: {f.name}: {e}")
            failed.append((f, str(e)))
        log("")

    log(f"Done: {len(ok)} succeeded, {len(failed)} failed.")
    return ok, failed


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "pdf_path",
        help="Path to a single Individual Clocking History PDF, OR a folder "
        "containing several such PDFs (batch mode - every *.pdf in the folder "
        "is parsed).",
    )
    ap.add_argument(
        "output_xlsx",
        nargs="?",
        default=None,
        help="Single-file mode: path to write the output .xlsx (default: "
        "<pdf name>_parsed.xlsx next to the input). Batch mode: an output "
        "folder to write all workbooks into (default: alongside each input PDF).",
    )
    ap.add_argument(
        "--work-days",
        default="mon,tue,wed,thu,fri",
        help="Comma-separated scheduled working days used to guess Shift/Planned "
        "(default: mon,tue,wed,thu,fri). Every other day is treated as OFF. "
        "There's no roster in the source PDF, so this is a guess, not a fact.",
    )
    ap.add_argument(
        "--hours-per-day",
        type=float,
        default=DEFAULT_HOURS_PER_DAY,
        help=f"Planned hours for each scheduled working day (default: {DEFAULT_HOURS_PER_DAY:g}).",
    )
    ap.add_argument(
        "--rotating",
        action="store_true",
        help="Use for shifts that don't follow a fixed weekly pattern (e.g. 4-on/4-off): "
        "ignores --work-days and instead treats any day with a clocking as a scheduled "
        "--hours-per-day day.",
    )
    args = ap.parse_args()

    try:
        work_days = parse_work_days(args.work_days)
    except ValueError as e:
        sys.exit(str(e))

    pdf_path = Path(args.pdf_path)
    if not pdf_path.exists():
        sys.exit(f"Path not found: {pdf_path}")

    if pdf_path.is_dir():
        out_dir = Path(args.output_xlsx) if args.output_xlsx else None
        ok, failed = run_batch(
            [pdf_path],
            output_dir=out_dir,
            work_days=work_days,
            hours_per_day=args.hours_per_day,
            rotating=args.rotating,
        )
        if not ok and not failed:
            sys.exit(f"No .pdf files found in {pdf_path}")
        if failed:
            print("Failed files:")
            for f, err in failed:
                print(f"  - {f}: {err}")
            sys.exit(1)
        return

    out_path = Path(args.output_xlsx) if args.output_xlsx else pdf_path.with_name(
        pdf_path.stem + "_parsed.xlsx"
    )
    process_one(
        pdf_path, out_path, work_days=work_days, hours_per_day=args.hours_per_day, rotating=args.rotating
    )


if __name__ == "__main__":
    main()
